"""
Excel ファイル操作モジュール
バーコードデータの管理と Excel 出力を担当
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from typing import List, Tuple


class ExcelHandler:
    """Excel ファイル操作クラス"""
    
    def __init__(self):
        """初期化"""
        self.barcodes = []
        self.duplicates = set()
    
    def add_barcode(self, barcode: str) -> Tuple[bool, str]:
        """
        バーコードをデータに追加
        
        Args:
            barcode: バーコード文字列
            
        Returns:
            (成功フラグ, メッセージ)
        """
        if barcode in self.barcodes:
            # 重複検出
            self.duplicates.add(barcode)
            return False, f"【重複】バーコード: {barcode} は既に登録されています（{self.barcodes.index(barcode) + 1}行目）"
        
        # 新規追加
        self.barcodes.append(barcode)
        return True, f"✓ バーコード追加完了: {barcode}"
    
    def get_barcodes(self) -> List[str]:
        """
        バーコード一覧を取得
        
        Returns:
            バーコードリスト
        """
        return self.barcodes.copy()
    
    def get_statistics(self) -> dict:
        """
        統計情報を取得
        
        Returns:
            統計情報（総件数、重複件数）
        """
        return {
            "total": len(self.barcodes),
            "duplicates": len(self.duplicates)
        }
    
    def clear_all(self) -> None:
        """すべてのデータをクリア"""
        self.barcodes.clear()
        self.duplicates.clear()
    
    def save_to_excel(self, file_path: str) -> Tuple[bool, str]:
        """
        Excel ファイルに保存
        
        Args:
            file_path: ファイルパス
            
        Returns:
            (成功フラグ, メッセージ)
        """
        try:
            # ワークブックを作成
            wb = Workbook()
            ws = wb.active
            ws.title = "Barcodes"
            
            # ヘッダーを設定
            headers = ["No.", "バーコード", "読み込み日時"]
            ws.append(headers)
            
            # ヘッダーのスタイル設定
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # データを追加
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            for idx, barcode in enumerate(self.barcodes, 1):
                ws.append([idx, barcode, now])
            
            # データのスタイル設定
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center_alignment = Alignment(horizontal="center", vertical="center")
            
            for row in ws.iter_rows(min_row=2, max_row=len(self.barcodes) + 1, min_col=1, max_col=3):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = center_alignment
            
            # 列幅を自動調整
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 20
            
            # ファイルを保存
            wb.save(file_path)
            
            return True, f"✓ Excel ファイルを保存しました: {file_path}"
        except Exception as e:
            return False, f"✗ 保存エラー: {str(e)}"
